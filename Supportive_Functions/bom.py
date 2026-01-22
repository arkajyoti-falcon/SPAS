# app.py
# pip install streamlit pandas openpyxl
# streamlit run app.py

import io
import os
import re
import math
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import openpyxl
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side


# -----------------------------
# Utilities
# -----------------------------
def norm(s: Any) -> str:
    if s is None:
        return ""
    return re.sub(r"[\s_\-]+", " ", str(s)).strip().lower()


def is_num(x: Any) -> bool:
    try:
        if x is None:
            return False
        # handle NaN
        if isinstance(x, float) and math.isnan(x):
            return False
        float(x)
        return True
    except Exception:
        return False


def to_float(x: Any, default: float = 0.0) -> float:
    try:
        if x is None:
            return default
        if isinstance(x, float) and math.isnan(x):
            return default
        if isinstance(x, str) and x.strip() == "":
            return default
        return float(x)
    except Exception:
        return default


def safe_int(x: Any, default: int = 0) -> int:
    """
    Safe int conversion for Excel cells:
    - handles None, "", NaN
    - handles numeric strings
    """
    try:
        if x is None:
            return default
        if isinstance(x, float) and math.isnan(x):
            return default
        if isinstance(x, str) and x.strip() == "":
            return default
        v = float(x)
        if math.isnan(v):
            return default
        return int(v)
    except Exception:
        return default


def fmt_len_m(x: float) -> str:
    if x <= 0:
        return ""
    s = f"{x:.2f}".rstrip("0").rstrip(".")
    return s


@dataclass
class ExtractedValue:
    value: Any
    source: str  # "Sheet!A1 (rule)"


# -----------------------------
# Sheet scoring (content-based)
# -----------------------------
ROLE_KEYWORDS = {
    "conveyors": ["conveyor length", "set", "name", "s no", "total conveyor"],
    "destinations": ["chute", "destinations", "qty", "description"],
    "ptl": ["ptl", "pick to light", "put to light", "modules", "control box"],
    "steelworks": ["steel", "platform", "sqm", "stairs", "area", "description", "qty"],
    "loop_cbs": ["loop cbs", "carrier pitch", "sorter", "pitch", "loop length", "speed", "drive"],
    "tech_specs": ["technical", "specifications", "speed", "feedline", "throughput"],
}

SHEETNAME_HINTS = {
    "conveyors": [r"(?i)\bconveyors?\b", r"(?i)\bconveyor\s+\b"],  # Case-insensitive: conveyor, conveyors, Conveyor, "Conveyor "
    "destinations": [r"\bdestinations?\b"],
    "ptl": [r"\bptl\b"],
    "steelworks": [r"\bsteelworks?\b", r"\bsteel works\b"],
    "loop_cbs": [r"\bloop cbs\b"],
    "tech_specs": [r"\btechnical\b", r"\bspecification\b"],
}


def sheet_text_sample(ws, max_rows=60, max_cols=20) -> str:
    parts = []
    rmax = min(ws.max_row or 1, max_rows)
    cmax = min(ws.max_column or 1, max_cols)
    for r in range(1, rmax + 1):
        for c in range(1, cmax + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                t = norm(v)
                if t:
                    parts.append(t)
    return " ".join(parts[:4000])


def score_sheet_for_role(ws, role: str) -> int:
    text = sheet_text_sample(ws)
    score = 0

    for kw in ROLE_KEYWORDS.get(role, []):
        if kw in text:
            score += 3

    name = norm(ws.title)
    for pat in SHEETNAME_HINTS.get(role, []):
        if re.search(pat, name):
            score += 5

    return score


def detect_sheets_by_role(wb) -> Dict[str, Any]:
    roles: Dict[str, Any] = {}
    scored = {role: [] for role in ROLE_KEYWORDS.keys()}

    for sname in wb.sheetnames:
        ws = wb[sname]
        for role in scored.keys():
            scored[role].append((score_sheet_for_role(ws, role), sname))

    for role, arr in scored.items():
        arr.sort(reverse=True, key=lambda x: x[0])
        best_score, best_name = arr[0]
        roles[role] = (best_name if best_score >= 6 else None)

    loop_candidates = []
    for sname in wb.sheetnames:
        if "loop cbs" in norm(sname):
            loop_candidates.append(sname)
    roles["loop_cbs_multi"] = loop_candidates

    return roles


# -----------------------------
# Generic table extraction by header detection
# -----------------------------
def find_header_row(ws, required: List[str], max_rows=80, max_cols=40) -> Optional[int]:
    req = [norm(x) for x in required]
    rmax = min(ws.max_row or 1, max_rows)
    cmax = min(ws.max_column or 1, max_cols)

    for r in range(1, rmax + 1):
        row_vals = [norm(ws.cell(r, c).value) for c in range(1, cmax + 1)]
        hit = 0
        for k in req:
            if any(k in v for v in row_vals):
                hit += 1
        if hit >= max(2, len(req) - 1):
            return r
    return None


def extract_table(ws, required_headers: List[str], stop_on: Optional[str] = "row labels") -> pd.DataFrame:
    hr = find_header_row(ws, required_headers)
    if hr is None:
        return pd.DataFrame()

    headers = {}
    cmax = min(ws.max_column or 1, 60)
    for c in range(1, cmax + 1):
        h = norm(ws.cell(hr, c).value)
        if h:
            headers[c] = h

    rows = []
    for r in range(hr + 1, (ws.max_row or hr + 1) + 1):
        row = {}
        empty = True
        for c, h in headers.items():
            v = ws.cell(r, c).value
            if isinstance(v, str) and stop_on and norm(v) == stop_on:
                return pd.DataFrame(rows)
            row[h] = v
            if v is not None and str(v).strip() != "":
                empty = False
        if empty:
            continue
        rows.append(row)

    return pd.DataFrame(rows)


# -----------------------------
# Label-value extraction
# -----------------------------
def find_value_near_label(
    ws,
    label_patterns: List[str],
    search_rows=220,
    search_cols=30,
    right_span=6,
    down_span=4,
) -> Optional[ExtractedValue]:
    pats = [re.compile(p, re.IGNORECASE) for p in label_patterns]
    rmax = min(ws.max_row or 1, search_rows)
    cmax = min(ws.max_column or 1, search_cols)

    for r in range(1, rmax + 1):
        for c in range(1, cmax + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            t = v.strip()
            if not t:
                continue
            if any(p.search(t) for p in pats):
                for dc in range(1, right_span + 1):
                    vv = ws.cell(r, c + dc).value
                    if vv is not None and str(vv).strip() != "":
                        return ExtractedValue(vv, f"{ws.title}!{openpyxl.utils.get_column_letter(c+dc)}{r} (right of '{t}')")
                for dr in range(1, down_span + 1):
                    vv = ws.cell(r + dr, c).value
                    if vv is not None and str(vv).strip() != "":
                        return ExtractedValue(vv, f"{ws.title}!{openpyxl.utils.get_column_letter(c)}{r+dr} (below '{t}')")
    return None


# -----------------------------
# Conveyors summarization
# -----------------------------
def normalize_conveyor_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    cols = {c: norm(c) for c in df.columns}
    df = df.rename(columns=cols)

    alias = {
        "conveyor length (mm)": "len_mm",
        "conveyor length(mm)": "len_mm",
        "length (mm)": "len_mm",
        "len (mm)": "len_mm",
        "conveyor length (m)": "len_m",
        "conveyor length(m)": "len_m",
        "length (m)": "len_m",
        "total conveyor length (m)": "total_m",
        "total conveyor length(m)": "total_m",
        "set": "set_qty",
        "qty": "set_qty",
        "name": "name",
        "family": "family",
    }

    out = {}
    for c in df.columns:
        nc = cols.get(c, norm(c))
        out[c] = alias.get(nc, nc)

    df = df.rename(columns=out)
    return df


def calc_row_total_m(row: pd.Series) -> float:
    total_m = to_float(row.get("total_m"), 0.0)
    if total_m > 0:
        return total_m

    set_qty = to_float(row.get("set_qty"), 0.0)
    mult = set_qty if set_qty > 0 else 1.0

    len_m = to_float(row.get("len_m"), 0.0)
    if len_m > 0:
        return len_m * mult

    len_mm = to_float(row.get("len_mm"), 0.0)
    if len_mm > 0:
        return (len_mm / 1000.0) * mult

    return 0.0


def count_units(group: pd.DataFrame) -> int:
    set_sum = safe_int(pd.to_numeric(group.get("set_qty", 0), errors="coerce").fillna(0).sum()) if "set_qty" in group.columns else 0

    if "len_mm" in group.columns:
        has_len = safe_int((pd.to_numeric(group["len_mm"], errors="coerce").fillna(0) > 0).sum())
    elif "len_m" in group.columns:
        has_len = safe_int((pd.to_numeric(group["len_m"], errors="coerce").fillna(0) > 0).sum())
    else:
        has_len = len(group)

    return max(set_sum, has_len, len(group))


def summarize_conveyors(df_conv: pd.DataFrame) -> pd.DataFrame:
    if df_conv.empty:
        return pd.DataFrame(columns=["name", "units", "total_m"])
    df = df_conv.copy()
    df["name"] = df.get("name", "").astype(str)
    df["name_l"] = df["name"].map(norm)
    df["total_m_calc"] = df.apply(calc_row_total_m, axis=1)

    out = []
    for name, g in df.groupby("name"):
        total_m = float(g["total_m_calc"].sum())
        units = count_units(g)
        out.append({"name": name, "units": units, "total_m": total_m})
    res = pd.DataFrame(out)
    res["name_l"] = res["name"].map(norm)
    return res.sort_values("name")


# -----------------------------
# Template detection
# -----------------------------
def detect_template(roles: Dict[str, Any], conv_sum: pd.DataFrame, loop_sheets: List[str]) -> str:
    nm = " ".join([norm(s) for s in loop_sheets])
    if ("upper" in nm and "lower" in nm) or ("upper deck" in nm and "lower deck" in nm):
        return "FM_DUAL_DECK"
    if roles.get("ptl"):
        return "NOON_POS"
    return "FM_SINGLE"


# -----------------------------
# BOM Builders
# -----------------------------
def add_row(rows: List[Dict[str, Any]], pos: Any, qty: Any, desc: str, val: str):
    rows.append({"Pos": pos, "Qty": qty, "Description": desc, "Value": val})


def build_noon_pos_bom(wb, roles: Dict[str, Any], conv_sum: pd.DataFrame, debug: List[str]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []

    loop_sheet = None
    for s in roles.get("loop_cbs_multi", []):
        if norm(s) == "loop cbs":
            loop_sheet = s
            break
    if not loop_sheet and roles.get("loop_cbs_multi"):
        loop_sheet = roles["loop_cbs_multi"][0]

    tech_sheet = roles.get("tech_specs")
    dest_sheet = roles.get("destinations")
    ptl_sheet = roles.get("ptl")

    zones = 1
    if loop_sheet:
        ws_loop = wb[loop_sheet]
        z = find_value_near_label(ws_loop, [r"induct\s*zones", r"number\s*of\s*zones"])
        zones = max(1, safe_int(z.value, 1)) if z else 1
        if z:
            debug.append(f"Zones={zones} from {z.source}")

    feedlines_total = ""
    if tech_sheet and loop_sheet:
        ws_tech = wb[tech_sheet]
        v = find_value_near_label(ws_tech, [r"feedlines?\s*in\s*zone", r"feedlines?\s*per\s*zone"])
        if v:
            feed_per_zone = to_float(v.value, 0.0)
            if feed_per_zone > 0:
                feedlines_total = safe_int(round(feed_per_zone * zones), 0)
                debug.append(f"Feedlines per zone={feed_per_zone} => total={feedlines_total} from {v.source}")

    def pick_by_name(patterns: List[str]) -> pd.DataFrame:
        pats = [re.compile(p, re.IGNORECASE) for p in patterns]
        return conv_sum[conv_sum["name"].astype(str).apply(lambda x: any(p.search(x) for p in pats))]

    pvc = pick_by_name([r"\bpvc\b", r"pvc belt conveyor"])
    modular = pick_by_name([r"\bmodular\b"])
    bag_takeaway = pick_by_name([r"\bbagging\b", r"bag takeaway", r"bagging conveyor"])

    pvc_len = float(pvc["total_m"].sum()) if not pvc.empty else 0.0
    pvc_units = safe_int(pvc["units"].sum(), 0) if not pvc.empty else 0
    modular_len = float(modular["total_m"].sum()) if not modular.empty else 0.0
    modular_units = safe_int(modular["units"].sum(), 0) if not modular.empty else 0
    bag_len = float(bag_takeaway["total_m"].sum()) if not bag_takeaway.empty else 0.0
    bag_units = safe_int(bag_takeaway["units"].sum(), 0) if not bag_takeaway.empty else 0

    sliding_per = rejection_per = nonsort_per = irr_chutes = 0
    if dest_sheet:
        ws_dest = wb[dest_sheet]
        df_dest = extract_table(ws_dest, ["description", "qty"])
        if not df_dest.empty:
            df_dest.columns = [norm(c) for c in df_dest.columns]
            desc_col = next((c for c in df_dest.columns if "description" in c), None)
            qty_col = next((c for c in df_dest.columns if c in ("qty", "quantity") or "qty" in c), None)

            if desc_col and qty_col:
                def qty_of(key: str) -> int:
                    m = df_dest[df_dest[desc_col].astype(str).str.lower().str.contains(key, na=False)]
                    if m.empty:
                        return 0
                    return safe_int(m.iloc[0][qty_col], 0)

                sliding_total = qty_of("sliding chute")
                rejection_total = qty_of("rejection")
                nonsort_total = qty_of("collection")
                irr_chutes = qty_of("irregular")

                sliding_per = safe_int(sliding_total / zones if zones else sliding_total, 0)
                rejection_per = safe_int(rejection_total / zones if zones else rejection_total, 0)
                nonsort_per = safe_int(nonsort_total / zones if zones else nonsort_total, 0)

    z1_bag = z2_bag = z1_pal = z2_pal = 0
    if ptl_sheet:
        ws_ptl = wb[ptl_sheet]
        df_ptl = extract_table(ws_ptl, ["description", "qty"])
        if not df_ptl.empty:
            df_ptl.columns = [norm(c) for c in df_ptl.columns]
            desc_col = next((c for c in df_ptl.columns if "description" in c), None)
            qty_col = next((c for c in df_ptl.columns if c in ("qty", "quantity") or "qty" in c), None)

            if desc_col and qty_col:
                def q(desc_key: str) -> int:
                    m = df_ptl[df_ptl[desc_col].astype(str).str.lower().str.contains(desc_key, na=False)]
                    if m.empty:
                        return 0
                    return safe_int(m.iloc[0][qty_col], 0)

                bagging_total = q("ptl racks bagging")
                pallets_total = q("ptl frame")

                if zones == 2:
                    z1_bag = bagging_total // 2
                    z2_bag = bagging_total - z1_bag
                    z1_pal = max(0, pallets_total // 2 - 1) if pallets_total else 0
                    z2_pal = pallets_total - z1_pal if pallets_total else 0
                else:
                    z1_bag = bagging_total
                    z1_pal = pallets_total

    height_mm = ""
    loop_len_txt = ""
    speed_txt = ""
    drive_txt = ""
    pitch_txt = ""

    if loop_sheet:
        ws_loop = wb[loop_sheet]
        h = find_value_near_label(ws_loop, [r"sorter\s*height", r"select.*sorter\s*height"])
        if h:
            hv = str(h.value)
            m = re.search(r"(\d{3,5})", hv)
            if m:
                height_mm = f"{m.group(1)} mm"
            elif "3" in hv:
                height_mm = "3300mm"
            debug.append(f"Height from {h.source} => {height_mm}")

        ll = find_value_near_label(ws_loop, [r"loop\s*length", r"sorter\s*length"])
        if ll and is_num(ll.value):
            loop_len_txt = f"~{safe_int(round(to_float(ll.value, 0.0)), 0)} m"
            debug.append(f"Loop length from {ll.source} => {loop_len_txt}")

        p = find_value_near_label(ws_loop, [r"carrier\s*pitch", r"\bpitch\b"])
        if p and is_num(p.value):
            pitch_txt = f"{safe_int(p.value, 0)} mm"
            debug.append(f"Pitch from {p.source} => {pitch_txt}")

        d = find_value_near_label(ws_loop, [r"drive", r"sorter\s*drive"])
        if d:
            dv = str(d.value).strip()
            drive_txt = "Linear Motor" if norm(dv) in ("lim", "linear", "linear motor") else dv
            debug.append(f"Drive from {d.source} => {drive_txt}")

    if tech_sheet:
        ws_tech = wb[tech_sheet]
        sp = find_value_near_label(ws_tech, [r"speed", r"sorter\s*speed"])
        if sp and is_num(sp.value):
            speed_txt = f"~{to_float(sp.value, 0.0):g} m/s"
            debug.append(f"Speed from {sp.source} => {speed_txt}")

    add_row(rows, 1, 1, "Infeed System", "")
    if pvc_len > 0:
        add_row(rows, "", "", f"• PVC belt Conveyor- {fmt_len_m(pvc_len)} m; {pvc_units} Modules", "Included")
    if modular_len > 0:
        add_row(rows, "", "", f"• Modular Conveyor- {fmt_len_m(modular_len)} m; {modular_units} Modules", "Included")
    add_row(rows, "", "", "• Includes Additional FC Conveyor", "Included")
    add_row(rows, "", "", "• 2 x XL Arm Diverters", "Included")
    add_row(rows, "", "", "", "")

    add_row(
        rows,
        2,
        feedlines_total,
        "Feedlines\nConsists of\n• Receiving Conveyor\n• Weighing Conveyor\n• Buffer Conveyor\n• Intelligent merge",
        "Included",
    )

    add_row(rows, 3, 1, "Irregular Shipment Handling (Optional)", "")
    add_row(rows, "", "", "• Powered Belt Conveyors", "Included")
    add_row(rows, "", "", "• Irregular chutes", f"{irr_chutes} Nos")
    add_row(rows, "", "", "", "")

    desc4 = (
        "Sorter\n"
        "Loop Cross Belt Sorter\n"
        "• Sorter Height\n"
        "• Sorter length\n"
        "• Sorter Speed\n"
        "• Drive\n"
        "• Pitch\n"
        "Including:\n"
        "• Standard Sorter Support Structure\n"
        "• Barcode scan system\n"
        "• Volume Detector\n"
        "• Automation control panel\n"
        "• Cable trays\n"
        "• Wiring & PLC\n"
        "• Safety net\n"
        "• Hooters\n"
        "• Fencing"
    )
    val4 = "\n".join([x for x in [height_mm, loop_len_txt, speed_txt, drive_txt, pitch_txt] if x])
    add_row(rows, 4, 1, desc4, val4)

    add_row(rows, 5, zones, "Sorter Outputs", "")
    add_row(rows, "", "", "• Sliding Chutes", f"{sliding_per} Nos")
    add_row(rows, "", "", "• Rejection chutes", f"{rejection_per:02d} Nos" if 0 < rejection_per < 10 else f"{rejection_per} Nos")
    add_row(rows, "", "", "• Non-Sort Chutes", f"{nonsort_per} Nos")

    desc6 = (
        "PTL for Zone 1\n"
        "• PTL for Bagging Type Rack\n"
        "• PTL for Palletizing Area\n"
        "PTL for Zone 2\n"
        "• PTL for Bagging\n"
        "• PTL for Pallets"
    )
    val6_lines = []
    if zones >= 1:
        val6_lines.append(f"{z1_bag} Nos")
        val6_lines.append(f"{z1_pal} Nos")
    if zones >= 2:
        val6_lines.append(f"{z2_bag} Nos")
        val6_lines.append(f"{z2_pal} Nos")
    add_row(rows, 6, 1, desc6, "\n".join(val6_lines))

    add_row(rows, 7, 1, "Bag Takeaway Conveyor", "")
    if bag_len > 0:
        add_row(rows, "", "", f"• Powered Belt Conveyors- {fmt_len_m(bag_len)} m; {bag_units} Modules", "Included")

    return pd.DataFrame(rows)


def classify_conveyor_groups(conv_sum: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
    groups: Dict[str, Dict[str, Any]] = {}

    def match(mask_pat: str) -> pd.DataFrame:
        p = re.compile(mask_pat, re.IGNORECASE)
        return conv_sum[conv_sum["name"].astype(str).apply(lambda x: bool(p.search(x)))]

    powered = conv_sum[
        conv_sum["name_l"].str.contains("pvc", na=False)
        & ~conv_sum["name_l"].str.contains("modular|turn|curve|telescopic|idler|merge|align|buffer|tbc", na=False)
    ]
    if powered.empty:
        powered = match(r"PVC\s*Belt\s*Conveyor|Powered\s*Belt")

    modular = match(r"\bmodular\b")
    merge = match(r"\bmerge\b")
    aligning = match(r"\balign")
    turn_90 = match(r"90[_\s-]*deg|90\s*degree|90\s*deg|turn.*90|curve.*90")
    turn_30 = match(r"30[_\s-]*deg|30\s*degree|30\s*deg|turn.*30|curve.*30")
    turn_60 = match(r"60[_\s-]*deg|60\s*degree|60\s*deg|turn.*60|curve.*60")
    sing = match(r"singulator")
    telescopic = match(r"telescopic|\btbc\b")
    idler = match(r"idler")
    buffer = match(r"buffer")

    def add_group(key: str, label: str, df: pd.DataFrame):
        if df is None or df.empty:
            return
        groups[key] = {
            "label": label,
            "total_m": float(df["total_m"].sum()),
            "units": safe_int(df["units"].sum(), 0),
        }

    add_group("powered", "Powered Belt Conveyors", powered)
    add_group("powered_optional", "Powered Belt Conveyors (Optional)",
              conv_sum[conv_sum["name_l"].str.contains("optional", na=False) & conv_sum["name_l"].str.contains("pvc|powered", na=False)])
    add_group("modular", "Modular Belt Conveyor", modular)
    add_group("merge", "Merge Conveyor", merge)
    add_group("merge_optional", "Merge Conveyor (Optional)",
              conv_sum[conv_sum["name_l"].str.contains("optional", na=False) & conv_sum["name_l"].str.contains("merge", na=False)])
    add_group("aligning", "Aligning Conveyor", aligning)
    add_group("buffer", "Buffer Conveyor", buffer)
    add_group("turn_90", "Curve Conveyor- 90 Degree", turn_90)
    add_group("turn_60", "Curve Conveyor- 60 Degree", turn_60)
    add_group("turn_30", "Curve Conveyor- 30 Degree", turn_30)
    add_group("sing", "Singulators", sing)
    add_group("telescopic", "Telescopic Belt Conveyor", telescopic)
    add_group("idler", "Idler Roller Conveyors", idler)

    return groups


def build_fm_bom(wb, roles: Dict[str, Any], conv_sum: pd.DataFrame, debug: List[str], dual_deck: bool) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []

    total_len = float(conv_sum["total_m"].sum()) if not conv_sum.empty else 0.0
    total_units = safe_int(conv_sum["units"].sum(), 0) if not conv_sum.empty else 0

    fm1_title = "Conveyor Package" if dual_deck else "Infeed System"
    fm1_desc = fm1_title
    if total_len > 0 and total_units > 0:
        fm1_desc = f"{fm1_title}\n~{fmt_len_m(total_len)} Metres ({total_units} Conveyors)"

    add_row(rows, "FM1", 1, fm1_desc, "")

    groups = classify_conveyor_groups(conv_sum)

    def bullet_line(label: str, total_m: float, units: int) -> Tuple[str, str]:
        if total_m > 0:
            return (f"• {label} ~{fmt_len_m(total_m)} Metres ({units} Conveyors)", "")
        return (f"• {label}", f"{units} Nos" if units > 0 else "")

    order = ["powered", "powered_optional", "modular", "merge", "merge_optional", "aligning",
             "turn_90", "turn_60", "turn_30", "sing", "telescopic", "idler", "buffer"]
    for k in order:
        g = groups.get(k)
        if not g:
            continue
        d, v = bullet_line(g["label"], g["total_m"], g["units"])
        add_row(rows, "", "", d, v)

    def infer_feedlines_from_sheets() -> List[Tuple[str, int, str]]:
        results = []
        for sname in wb.sheetnames:
            if any(x in norm(sname) for x in ["induct", "loop cbs", "linear cbs"]):
                ws = wb[sname]
                v = find_value_near_label(ws, [r"total\s*inductions?", r"total\s*inducts?"])
                if v and is_num(v.value):
                    inductions = safe_int(v.value, 0)
                    feedlines = inductions * 2
                    results.append((sname, feedlines, v.source))
        return results

    feed_candidates = infer_feedlines_from_sheets()
    if feed_candidates:
        feed_candidates.sort(key=lambda x: x[1], reverse=True)
        used = []
        for sname, cnt, src in feed_candidates:
            if cnt not in [u[1] for u in used]:
                used.append((sname, cnt, src))
            if dual_deck and len(used) >= 2:
                break
        if dual_deck and len(used) >= 2:
            add_row(rows, "FM2", used[0][1], "Manual Induct Feedlines", "Included")
            debug.append(f"FM2 feedlines={used[0][1]} from {used[0][2]}")
            add_row(rows, "FM3", used[1][1], "Auto Induct Feedlines", "Included")
            debug.append(f"FM3 feedlines={used[1][1]} from {used[1][2]}")
        else:
            add_row(rows, "FM2", used[0][1], "Feedlines", "Included")
            debug.append(f"FM2 feedlines={used[0][1]} from {used[0][2]}")
    else:
        tech_sheet = roles.get("tech_specs")
        loop_sheets = roles.get("loop_cbs_multi", [])
        zones = 1
        if loop_sheets:
            ws_loop = wb[loop_sheets[0]]
            z = find_value_near_label(ws_loop, [r"induct\s*zones", r"number\s*of\s*zones"])
            zones = max(1, safe_int(z.value, 1)) if z else 1

        if tech_sheet:
            ws_tech = wb[tech_sheet]
            v = find_value_near_label(ws_tech, [r"feedlines?\s*in\s*zone", r"feedlines?\s*per\s*zone"])
            if v and is_num(v.value):
                total = safe_int(round(to_float(v.value, 0.0) * zones), 0)
                add_row(rows, "FM2", total, "Feedlines", "Included")
                debug.append(f"FM2 feedlines={total} from {v.source}")

    loop_sheets = roles.get("loop_cbs_multi", [])
    chosen_loops = []

    if dual_deck:
        lower = next((s for s in loop_sheets if "lower" in norm(s)), None)
        upper = next((s for s in loop_sheets if "upper" in norm(s)), None)
        if lower:
            chosen_loops.append(("FM4", "Falcon Lower Deck Sorter", lower))
        if upper:
            chosen_loops.append(("FM5", "Falcon Upper Deck Sorter", upper))
        if len(chosen_loops) == 0 and len(loop_sheets) >= 1:
            chosen_loops.append(("FM4", "Falcon Sorter", loop_sheets[0]))
        if len(chosen_loops) == 1 and len(loop_sheets) >= 2:
            chosen_loops.append(("FM5", "Falcon Sorter", loop_sheets[1]))
    else:
        if loop_sheets:
            chosen_loops.append(("FM3", "Sorter", loop_sheets[0]))

    tech_sheet = roles.get("tech_specs")

    def extract_sorter_params(ws_loop) -> Dict[str, str]:
        height = ""
        loop_len = ""
        speed = ""
        drive = ""
        pitch = ""

        h = find_value_near_label(ws_loop, [r"sorter\s*height", r"select.*sorter\s*height"])
        if h:
            hv = str(h.value)
            m = re.search(r"(\d{3,5})", hv)
            if m:
                height = f"{m.group(1)} mm"
            elif "3" in hv:
                height = "3300mm"

        ll = find_value_near_label(ws_loop, [r"loop\s*length", r"sorter\s*length"])
        if ll and is_num(ll.value):
            loop_len = f"~{safe_int(round(to_float(ll.value, 0.0)), 0)} m"

        p = find_value_near_label(ws_loop, [r"carrier\s*pitch", r"\bpitch\b"])
        if p and is_num(p.value):
            pitch = f"{safe_int(p.value, 0)} mm"

        d = find_value_near_label(ws_loop, [r"drive", r"sorter\s*drive"])
        if d:
            dv = str(d.value).strip()
            drive = "Linear Motor" if norm(dv) in ("lim", "linear", "linear motor") else dv

        if tech_sheet:
            ws_tech = wb[tech_sheet]
            sp = find_value_near_label(ws_tech, [r"speed", r"sorter\s*speed"])
            if sp and is_num(sp.value):
                speed = f"~{to_float(sp.value, 0.0):g} m/s"

        return {"height": height, "length": loop_len, "speed": speed, "drive": drive, "pitch": pitch}

    sorter_common_desc = (
        "1 Loop Cross Belt Sorter\n"
        "• Sorter Height\n"
        "• Sorter Length\n"
        "• Sorter Speed\n"
        "• Sorter Drive\n"
        "• Carrier Pitch\n"
        "Including:\n"
        "• Standard Sorter Supports\n"
        "• Empty Carrier Detection System\n"
        "• Product Centring System\n"
        "• Dimension Scanning System\n"
        "• E-Stops\n"
        "• Hooters\n"
        "• Fencing\n"
        "• Netting\n"
        "• Overshoot Assembly"
    )

    for fm_pos, title, sheet in chosen_loops:
        ws_loop = wb[sheet]
        params = extract_sorter_params(ws_loop)
        val = "\n".join([x for x in [params["height"], params["length"], params["speed"], params["drive"], params["pitch"]] if x])
        add_row(rows, fm_pos, 1, f"{title}\n{sorter_common_desc}", val)

    dest_sheet = roles.get("destinations")
    if dest_sheet:
        ws_dest = wb[dest_sheet]
        df_dest = extract_table(ws_dest, ["description", "qty"])
        if not df_dest.empty:
            df_dest.columns = [norm(c) for c in df_dest.columns]
            desc_col = next((c for c in df_dest.columns if "description" in c), None)
            qty_col = next((c for c in df_dest.columns if "qty" in c or c in ("qty", "quantity")), None)

            add_row(rows, "FM6" if dual_deck else "FM4", 1, "Sorter Outputs", "")
            if desc_col and qty_col:
                for _, r in df_dest.iterrows():
                    d = str(r.get(desc_col, "")).strip()
                    if not d:
                        continue
                    q = r.get(qty_col, "")
                    qtxt = f"{safe_int(q, 0)} Nos" if is_num(q) else str(q)
                    add_row(rows, "", "", f"• {d}", qtxt)

    steel_sheet = roles.get("steelworks")
    if steel_sheet:
        ws_steel = wb[steel_sheet]
        df_sw = extract_table(ws_steel, ["description", "qty"])
        if not df_sw.empty:
            df_sw.columns = [norm(c) for c in df_sw.columns]
            desc_col = next((c for c in df_sw.columns if "description" in c), None)
            qty_col = next((c for c in df_sw.columns if "qty" in c or c in ("qty", "quantity")), None)
            area_col = next((c for c in df_sw.columns if "area" in c), None)

            add_row(rows, "FM7" if dual_deck else "FM5", 1, "Steel Works", "")
            if desc_col:
                for _, r in df_sw.iterrows():
                    d = str(r.get(desc_col, "")).strip()
                    if not d:
                        continue
                    val = ""
                    if area_col and is_num(r.get(area_col)):
                        val = f"~{fmt_len_m(to_float(r.get(area_col), 0.0))} SQM"
                    elif qty_col and is_num(r.get(qty_col)):
                        val = f"{safe_int(r.get(qty_col), 0)} Nos"
                    add_row(rows, "", "", f"• {d}", val)

    fire_door_qty = None
    for sname in wb.sheetnames:
        ws = wb[sname]
        v = find_value_near_label(ws, [r"fire\s*door"])
        if v and is_num(v.value):
            vv = to_float(v.value, 0.0)
            if 0 < vv <= 50:
                fire_door_qty = safe_int(vv, 0)
                debug.append(f"Fire Door qty={fire_door_qty} from {v.source}")
                break
    if fire_door_qty is not None:
        add_row(rows, "FM8" if dual_deck else "FM6", 1, "Fire Door", f"{fire_door_qty} Nos")

    return pd.DataFrame(rows)


# -----------------------------
# Excel export
# -----------------------------
def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Mechanical_BOM") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="top", wrap_text=True)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = list(df.columns)
    for c, col in enumerate(cols, start=1):
        cell = ws.cell(1, c, value=col)
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, value=val)
            cell.alignment = wrap if cols[c_idx - 1] in ("Description", "Value") else center
            cell.border = border

    width_map = {"Pos": 10, "Qty": 8, "Description": 75, "Value": 30}
    for i, col in enumerate(cols, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width_map.get(col, 18)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
