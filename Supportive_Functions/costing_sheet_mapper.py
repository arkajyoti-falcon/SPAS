"""
Costing Sheet Mapper - Extract and map component data from Excel costing sheets

This module provides functionality to:
1. Detect available sheets in a costing workbook by component type
2. Extract component quantities from costing/BOQ sheets
3. Map component types to their canonical sheet names
4. Extract full sheet data to send to AI for fact building

Usage:
    mapper = CostingSheetMapper(workbook_path)
    loop_cbs_counts = mapper.extract_component_counts("Loop CBS")
    conveyors_data = mapper.extract_sheet_as_text("Conveyors")
"""

import re
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class SheetMapping:
    """Represents a discovered sheet and its detected role"""
    sheet_name: str
    component_type: str  # "Loop CBS", "Conveyors", "Destinations", etc.
    confidence: float  # 0.0 to 1.0
    detected_columns: List[str] = None


# Sheet name patterns for component type detection
COMPONENT_TYPE_PATTERNS = {
    "Loop CBS": [
        r"loop\s*cbs",
        r"loop",
        r"cbs.*loop",
    ],
    "Linear CBS": [
        r"linear\s*cbs",
        r"linear",
        r"cbs.*linear",
    ],
    "Conveyors": [
        r"conveyor[s]?",
        r"conveyor\s*boq",
        r"conveyor\s*list",
    ],
    "Destinations": [
        r"destination[s]?",
        r"chutes",
        r"output",
    ],
    "Steelworks": [
        r"steel\s*work[s]?",
        r"steelwork",
        r"platform",
        r"structure",
    ],
    "PTL": [
        r"\bptl\b",
        r"pick\s*to\s*light",
        r"put\s*to\s*light",
    ],
    "Induct Lines": [
        r"induct",
        r"feedline[s]?",
        r"induction",
    ],
    "Control System": [
        r"control",
        r"wcs",
        r"software",
    ],
    "Safety Equipment": [
        r"safety",
        r"guard",
        r"fencing",
    ],
}

# Standard column names to look for in each sheet type
EXPECTED_COLUMNS = {
    "Loop CBS": ["loop length", "carrier pitch", "speed", "drive", "motor"],
    "Linear CBS": ["length", "carriers", "carrier pitch", "speed"],
    "Conveyors": ["name", "conveyor length", "width", "set", "el_1", "el_2"],
    "Destinations": ["chute", "description", "qty", "quantity"],
    "Steelworks": ["description", "area", "sqm", "qty"],
    "Induct Lines": ["feedline", "induct", "station", "qty"],
}


def normalize_text(text: str) -> str:
    """Normalize text for comparison"""
    if not text:
        return ""
    return re.sub(r"[\s_\-]+", " ", str(text)).strip().lower()


def load_component_sheets(xlsx_path: str) -> Dict[str, Dict[str, Any]]:
    """
    Load all sheets from Excel workbook and build component registry.
    
    Normalizes sheet names and extracts key values from each sheet.
    
    Args:
        xlsx_path: Path to Excel workbook
        
    Returns:
        Dictionary mapping normalized_sheet_name to:
        {
            "sheet_name": original_sheet_name,
            "table": dataframe_or_rows,
            "key_values": { extracted important fields }
        }
        
    Example:
        registry = load_component_sheets("costing.xlsx")
        # registry["weighing conveyors"]["table"] -> DataFrame
        # registry["weighing conveyors"]["key_values"] -> {"quantity": 5, ...}
    """
    import pandas as pd
    
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        raise ValueError(f"Failed to load workbook {xlsx_path}: {e}")
    
    registry = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Normalize sheet name: lowercase, replace underscores/hyphens with space, collapse spaces
        normalized_name = normalize_text(sheet_name)
        
        # Extract table as dataframe
        try:
            # Read the sheet using pandas, trying to auto-detect header
            df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=None)
            
            # Find header row (first row with mostly non-empty values)
            header_row = None
            for idx in range(min(20, len(df))):
                non_empty = df.iloc[idx].notna().sum()
                if non_empty >= len(df.columns) * 0.5:  # At least 50% non-empty
                    header_row = idx
                    break
            
            if header_row is not None and header_row > 0:
                # Re-read with detected header
                df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=header_row)
            elif header_row == 0:
                # First row is header, re-read to set it properly
                df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=0)
            
            # Clean up the dataframe
            df = df.dropna(how='all')  # Remove all-empty rows
            
        except Exception as e:
            # Fallback: read as raw rows
            rows = []
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    rows.append(row)
            df = rows
        
        # Extract key values from first few rows and columns
        key_values = _extract_key_values_from_sheet(ws, sheet_name)
        
        registry[normalized_name] = {
            "sheet_name": sheet_name,
            "table": df,
            "key_values": key_values
        }
    
    return registry


def _extract_key_values_from_sheet(ws: Worksheet, sheet_name: str) -> Dict[str, Any]:
    """
    Extract important field values from a worksheet.
    
    Looks for common patterns like "Quantity: 5", "Total: 100", etc.
    
    Args:
        ws: openpyxl Worksheet
        sheet_name: Name of the sheet (for context)
        
    Returns:
        Dictionary of extracted key-value pairs
    """
    key_values = {}
    
    # Common patterns to look for
    patterns = {
        "quantity": [r"qty", r"quantity", r"count", r"total\s+qty"],
        "length": [r"length", r"total\s+length"],
        "width": [r"width", r"carrier"],
        "speed": [r"speed\s*\("],
        "pitch": [r"pitch"],
        "total": [r"total", r"sum"],
        "diameter": [r"diameter", r"dia"],
        "capacity": [r"capacity"],
    }
    
    # Scan first 50 rows and columns for key-value pairs
    for row_idx in range(min(50, ws.max_row + 1)):
        for col_idx in range(1, min(10, ws.max_column + 1)):
            cell = ws.cell(row_idx, col_idx)
            cell_value = cell.value
            
            if not cell_value or not isinstance(cell_value, str):
                continue
            
            cell_norm = normalize_text(cell_value)
            
            # Check if this cell matches any pattern
            for key, patterns_list in patterns.items():
                for pattern in patterns_list:
                    if re.search(pattern, cell_norm, re.IGNORECASE):
                        # Look at next cell for the value
                        next_cell = ws.cell(row_idx, col_idx + 1)
                        next_value = next_cell.value
                        
                        if next_value is not None:
                            try:
                                # Try to extract numeric value
                                if isinstance(next_value, (int, float)):
                                    key_values[key] = next_value
                                elif isinstance(next_value, str):
                                    # Try to extract number from string
                                    num_match = re.search(r"[-+]?(\d+\.?\d*)", next_value)
                                    if num_match:
                                        try:
                                            key_values[key] = float(num_match.group(0))
                                        except ValueError:
                                            key_values[key] = next_value
                                    else:
                                        key_values[key] = next_value
                            except Exception:
                                pass
    
    return key_values


class CostingSheetMapper:
    """Main class for mapping and extracting component data from costing sheets"""
    
    def __init__(self, workbook_path: str):
        """
        Initialize mapper with a costing workbook.
        
        Args:
            workbook_path: Path to Excel workbook
        """
        self.workbook_path = Path(workbook_path)
        self.workbook = None
        self.sheet_mappings: Dict[str, SheetMapping] = {}
        
        self._load_workbook()
        self._detect_sheet_types()
    
    def _load_workbook(self):
        """Load the Excel workbook"""
        try:
            self.workbook = openpyxl.load_workbook(
                self.workbook_path,
                data_only=True,
                read_only=False
            )
        except Exception as e:
            raise ValueError(f"Failed to load workbook: {e}")
    
    def _detect_sheet_types(self):
        """Detect the type of each sheet in the workbook"""
        if not self.workbook:
            return
        
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            
            # Get sheet name and first 100 rows of data for analysis
            sheet_text = self._extract_sheet_sample(ws, max_rows=100)
            sheet_text_normalized = normalize_text(sheet_text)
            
            # Score each component type
            best_match = None
            best_score = 0.0
            
            for component_type, patterns in COMPONENT_TYPE_PATTERNS.items():
                score = 0.0
                
                # Check pattern matches in sheet name
                sheet_name_norm = normalize_text(sheet_name)
                for pattern in patterns:
                    if re.search(pattern, sheet_name_norm, re.IGNORECASE):
                        score += 0.5
                
                # Check pattern matches in content
                for pattern in patterns:
                    if re.search(pattern, sheet_text_normalized, re.IGNORECASE):
                        score += 0.3
                
                # Check for expected columns
                if component_type in EXPECTED_COLUMNS:
                    found_columns = 0
                    for expected_col in EXPECTED_COLUMNS[component_type]:
                        if expected_col in sheet_text_normalized:
                            found_columns += 1
                    if EXPECTED_COLUMNS[component_type]:
                        score += (found_columns / len(EXPECTED_COLUMNS[component_type])) * 0.2
                
                if score > best_score:
                    best_score = score
                    best_match = component_type
            
            if best_match and best_score > 0.1:
                detected_cols = self._extract_headers(ws)
                self.sheet_mappings[sheet_name] = SheetMapping(
                    sheet_name=sheet_name,
                    component_type=best_match,
                    confidence=min(best_score, 1.0),
                    detected_columns=detected_cols
                )
    
    def _extract_sheet_sample(self, ws: Worksheet, max_rows: int = 100) -> str:
        """Extract text sample from worksheet for analysis"""
        parts = []
        for row in ws.iter_rows(min_row=1, max_row=max_rows, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    parts.append(cell.strip())
                elif cell and not isinstance(cell, (int, float)):
                    parts.append(str(cell).strip())
        return " ".join(parts[:5000])  # Limit to 5000 chars
    
    def _extract_headers(self, ws: Worksheet) -> List[str]:
        """Extract column headers from worksheet"""
        headers = []
        
        # Check first 20 rows for header row
        for row_idx in range(1, min(21, ws.max_row + 1)):
            potential_headers = []
            non_empty = 0
            
            for col_idx in range(1, min(21, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                if cell.value:
                    potential_headers.append(normalize_text(str(cell.value)))
                    non_empty += 1
            
            # If we found a decent number of headers, use this row
            if non_empty >= 3 and len(potential_headers) > 0:
                return potential_headers
        
        return headers
    
    def get_sheet_by_component_type(self, component_type: str) -> Optional[str]:
        """
        Get the sheet name for a specific component type.
        
        Args:
            component_type: e.g., "Loop CBS", "Conveyors", "Destinations"
        
        Returns:
            Sheet name if found, else None
        """
        best_match = None
        best_confidence = 0.0
        
        for sheet_name, mapping in self.sheet_mappings.items():
            if mapping.component_type == component_type:
                if mapping.confidence > best_confidence:
                    best_confidence = mapping.confidence
                    best_match = sheet_name
        
        return best_match
    
    def get_all_sheet_mappings(self) -> Dict[str, SheetMapping]:
        """Get all detected sheet mappings"""
        return self.sheet_mappings
    
    def extract_sheet_as_text(self, sheet_identifier: str, max_rows: Optional[int] = None) -> str:
        """
        Extract entire sheet content as formatted text for AI processing.
        
        Args:
            sheet_identifier: Either sheet name or component type
            max_rows: Max rows to extract (None = all)
        
        Returns:
            Formatted text representation of sheet
        """
        # Resolve sheet name
        sheet_name = sheet_identifier
        
        # Try to find sheet by component type if not found by name
        if sheet_name not in self.workbook.sheetnames:
            sheet_name = self.get_sheet_by_component_type(sheet_identifier)
            if not sheet_name:
                raise ValueError(f"Sheet not found: {sheet_identifier}")
        
        ws = self.workbook[sheet_name]
        lines = [f"Sheet: {sheet_name}"]
        lines.append("=" * 80)
        
        max_r = max_rows if max_rows else ws.max_row
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_r, values_only=True), 1):
            row_parts = []
            for cell_val in row:
                if cell_val is None:
                    row_parts.append("")
                else:
                    row_parts.append(str(cell_val).strip())
            
            # Skip completely empty rows
            if any(p for p in row_parts):
                lines.append(" | ".join(row_parts))
        
        return "\n".join(lines)
    
    def extract_component_counts(self, component_type: str) -> Dict[str, int]:
        """
        Extract component quantities from a specific sheet.
        
        Args:
            component_type: e.g., "Loop CBS", "Conveyors"
        
        Returns:
            Dictionary of component names -> quantities
        """
        sheet_name = self.get_sheet_by_component_type(component_type)
        if not sheet_name:
            return {}
        
        ws = self.workbook[sheet_name]
        counts = {}
        
        # Find quantity column and name column
        qty_col = None
        name_col = None
        desc_col = None
        
        # Search for header row
        for row_idx in range(1, min(20, ws.max_row + 1)):
            for col_idx in range(1, ws.max_column + 1):
                cell_val = normalize_text(ws.cell(row_idx, col_idx).value)
                
                if "qty" in cell_val or "quantity" in cell_val or "count" in cell_val:
                    qty_col = col_idx
                elif "name" in cell_val or "description" in cell_val:
                    if not name_col:
                        name_col = col_idx
                    else:
                        desc_col = col_idx
        
        # If we found a qty column, extract data
        if qty_col:
            start_row = 1
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row_idx, qty_col)
                if isinstance(cell.value, (int, float)) and cell.value > 0:
                    start_row = row_idx + 1
                    break
            
            # Extract quantities
            for row_idx in range(start_row, ws.max_row + 1):
                qty_cell = ws.cell(row_idx, qty_col)
                
                if qty_cell.value is None:
                    continue
                
                try:
                    qty = int(qty_cell.value)
                    
                    if qty > 0 and name_col:
                        name_cell = ws.cell(row_idx, name_col)
                        if name_cell.value:
                            name = normalize_text(name_cell.value)
                            if name:
                                counts[name] = qty
                except (ValueError, TypeError):
                    continue
        
        return counts
    
    def extract_table_with_headers(self, sheet_identifier: str) -> Tuple[List[str], List[List[Any]]]:
        """
        Extract a table with headers from a sheet.
        
        Args:
            sheet_identifier: Sheet name or component type
        
        Returns:
            (headers, data_rows)
        """
        sheet_name = sheet_identifier
        if sheet_name not in self.workbook.sheetnames:
            sheet_name = self.get_sheet_by_component_type(sheet_identifier)
            if not sheet_name:
                raise ValueError(f"Sheet not found: {sheet_identifier}")
        
        ws = self.workbook[sheet_name]
        
        headers = []
        header_row = None
        
        # Find header row
        for row_idx in range(1, min(30, ws.max_row + 1)):
            row_vals = []
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row_idx, col_idx).value
                row_vals.append(val)
            
            # Check if this looks like a header row
            if any(isinstance(v, str) for v in row_vals):
                headers = row_vals
                header_row = row_idx + 1
                break
        
        # Extract data rows
        data = []
        if header_row:
            for row_idx in range(header_row, ws.max_row + 1):
                row_vals = []
                has_data = False
                for col_idx in range(1, len(headers) + 1):
                    val = ws.cell(row_idx, col_idx).value
                    row_vals.append(val)
                    if val is not None:
                        has_data = True
                
                if has_data:
                    data.append(row_vals)
        
        return headers, data


# ============================================================================
# Convenience Functions
# ============================================================================

def extract_all_sheet_data(costing_file_path: str) -> Dict[str, Dict[str, Any]]:
    """
    Extract and summarize all available data from a costing workbook.
    
    Returns dict with:
    - sheet_name -> { type, confidence, headers, sample_data }
    """
    try:
        mapper = CostingSheetMapper(costing_file_path)
    except Exception as e:
        return {"error": str(e)}
    
    result = {}
    
    for sheet_name, mapping in mapper.get_all_sheet_mappings().items():
        try:
            headers, data = mapper.extract_table_with_headers(sheet_name)
            result[sheet_name] = {
                "component_type": mapping.component_type,
                "confidence": mapping.confidence,
                "columns": [str(h) for h in headers],
                "row_count": len(data),
                "sample_data": data[:5] if data else [],
            }
        except Exception as e:
            result[sheet_name] = {
                "component_type": mapping.component_type,
                "confidence": mapping.confidence,
                "error": str(e),
            }
    
    return result


if __name__ == "__main__":
    # Example usage
    import json
    
    # Test with sample file
    test_file = "F24-00276-BOSTA CAIRO_Loop CBS_Rev-11.xlsx"
    
    try:
        mapper = CostingSheetMapper(test_file)
        
        print("Detected Sheet Mappings:")
        for sheet_name, mapping in mapper.get_all_sheet_mappings().items():
            print(f"  {sheet_name}: {mapping.component_type} (confidence: {mapping.confidence:.2f})")
        
        print("\n" + "=" * 80)
        print("Loop CBS Sheet Content:")
        print("=" * 80)
        
        loop_cbs_data = mapper.extract_sheet_as_text("Loop CBS", max_rows=30)
        print(loop_cbs_data)
        
    except Exception as e:
        print(f"Error: {e}")
